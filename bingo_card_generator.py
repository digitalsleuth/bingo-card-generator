#!/usr/bin/env python3

'''
Requires wkhtmltopdf for OS it's being run on
(Windows requires an exe, linux install from pkg manager)

While not normally necessary for CSS, the float values are required for
the proper printing of the PDF's with wkhtmltopdf.
If these are removed, you can expect the PDF's to be off-center or misaligned.

CSS Maple Leaf author Andre Lopes - https://codepen.io/alldrops/pen/jAzZmw
'''

import argparse
from random import Random
import re
import base64
import csv
import imghdr
import os
import sys
import pdfkit
import webcolors
from PyQt6 import QtCore, QtGui, QtWidgets
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

__author__ = 'Corey Forman'
__date__ = '19 Feb 2023'
__version__ = '6.0.1'
__description__ = f'Bingo Card Generator {__version__}'
__source__ = 'https://github.com/digitalsleuth/bingo-card-generator'
__colour_groups__ = 'https://www.w3schools.com/colors/colors_groups.asp'


class UiDialog():
    """This is the main class for setting up the UI"""
    def setup_ui(self, Dialog):
        """Instantiate and layout dialog options"""
        Dialog.setObjectName("Dialog")
        Dialog.setFixedSize(390, 260)
        label_font = QtGui.QFont()
        label_font.setPointSize(10.5)
        label_font.setFamily("Arial")
        label_font.StyleHint("SansSerif")
        self.dauber_shape = QtWidgets.QComboBox(Dialog)
        self.dauber_shape.setGeometry(QtCore.QRect(134, 167, 131, 27))
        self.dauber_shape.setEditable(False)
        self.dauber_shape.setObjectName("dauber_shape")
        item_count = 1
        while item_count < 12:
            self.dauber_shape.addItem("")
            item_count += 1
        self.dauber_shape.currentIndexChanged[int].connect(self.on_select)
        self.dauber_shape_label = QtWidgets.QLabel(Dialog)
        self.dauber_shape_label.setGeometry(QtCore.QRect(10, 172, 121, 21))
        self.dauber_shape_label.setToolTipDuration(-1)
        self.dauber_shape_label.setObjectName("dauber_shape_label")
        self.dauber_shape_label.setFont(label_font)
        self.title_label = QtWidgets.QLabel(Dialog)
        self.title_label.setGeometry(QtCore.QRect(0, 21, 388, 20))
        title_font = QtGui.QFont()
        title_font.setWeight(700)
        title_font.setPointSize(11)
        title_font.setFamily("Arial")
        title_font.StyleHint("SansSerif")
        self.title_label.setFont(title_font)
        self.title_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.title_label.setObjectName("title_label")
        self.dauber_colour_label = QtWidgets.QLabel(Dialog)
        self.dauber_colour_label.setGeometry(QtCore.QRect(10, 132, 121, 21))
        self.dauber_colour_label.setToolTipDuration(-1)
        self.dauber_colour_label.setObjectName("dauber_colour_label")
        self.dauber_colour_label.setFont(label_font)
        self.card_colour_label = QtWidgets.QLabel(Dialog)
        self.card_colour_label.setGeometry(QtCore.QRect(10, 92, 121, 21))
        self.card_colour_label.setToolTipDuration(-1)
        self.card_colour_label.setObjectName("card_colour_label")
        self.card_colour_label.setFont(label_font)
        self.select_logo_button = QtWidgets.QPushButton(Dialog)
        self.select_logo_button.setGeometry(QtCore.QRect(280, 167, 100, 27))
        self.select_logo_button.setDefault(False)
        self.select_logo_button.setObjectName("select_logo_button")
        self.select_logo_button.clicked.connect(self.select_logo)
        self.select_logo_button.setEnabled(False)
        self.select_logo_button.setVisible(False)
        self.select_logo_button.setFont(label_font)
        self.select_result = QtWidgets.QLineEdit(Dialog)
        self.select_result.setObjectName("select_result")
        self.select_result.setReadOnly(True)
        self.select_result.setVisible(False)
        self.select_result.setGeometry(QtCore.QRect(8, 205, 372, 30))
        self.select_result.setFont(label_font)
        self.set_card_title = QtWidgets.QPushButton("Card Title", Dialog)
        self.set_card_title.setObjectName("set_card_title")
        self.set_card_title.setFont(label_font)
        self.set_card_title.clicked.connect(self.enter_title)
        self.set_card_title.setGeometry(QtCore.QRect(280, 47, 100, 31))
        self.card_title = QtWidgets.QLineEdit(Dialog)
        self.card_title.setVisible(False)
        self.close = QtWidgets.QPushButton(Dialog)
        self.close.setGeometry(QtCore.QRect(280, 127, 100, 31))
        self.close.setObjectName("close")
        self.close.setFont(label_font)
        self.close.clicked.connect(QtWidgets.QApplication.instance().quit)
        self.number_label = QtWidgets.QLabel(Dialog)
        self.number_label.setGeometry(QtCore.QRect(10, 52, 121, 21))
        self.number_label.setToolTipDuration(-1)
        self.number_label.setObjectName("number_label")
        self.number_label.setFont(label_font)
        self.number = QtWidgets.QLineEdit(Dialog)
        self.number.setGeometry(QtCore.QRect(134, 47, 131, 31))
        self.number.setText("1")
        self.number.setObjectName("number")
        self.generate = QtWidgets.QPushButton(Dialog)
        self.generate.setGeometry(QtCore.QRect(280, 87, 100, 31))
        self.generate.setDefault(False)
        self.generate.setFont(label_font)
        self.generate.setObjectName("generate")
        self.generate.clicked.connect(lambda: gui_everything(int(self.number.text()),
                                                             self.card_colour.text(),
                                                             self.dauber_colour.text(),
                                                             self.dauber_shape.currentText(),
                                                             self.get_directory(),
                                                             self.select_logo(),
                                                             self.allow_select(),
                                                             self.easy_mode(),
                                                             self.card_title.text()))
        self.card_colour = QtWidgets.QLineEdit(Dialog)
        self.card_colour.setGeometry(QtCore.QRect(134, 87, 131, 31))
        self.card_colour.setObjectName("card_colour")
        self.card_colour_picker = QtWidgets.QPushButton(Dialog)
        self.card_colour_picker.setObjectName("card_colour_picker")
        self.card_colour_picker.clicked.connect(self.card_colourpicker)
        self.card_colour_picker.setGeometry(QtCore.QRect(110, 92, 20, 20))
        self.card_colour_picker.setStyleSheet("background-color: blue; border: 1px solid black")
        self.dauber_colour = QtWidgets.QLineEdit(Dialog)
        self.dauber_colour.setGeometry(QtCore.QRect(134, 127, 131, 31))
        self.dauber_colour.setObjectName("dauber_colour")
        self.dauber_colour.setEnabled(False)
        self.dauber_colour_picker = QtWidgets.QPushButton(Dialog)
        self.dauber_colour_picker.setObjectName("dauber_colour_picker")
        self.dauber_colour_picker.clicked.connect(self.dauber_colourpicker)
        self.dauber_colour_picker.setGeometry(QtCore.QRect(110, 132, 20, 20))
        self.dauber_colour_picker.setStyleSheet("background-color: red; border: 1px solid black")
        self.dauber_colour_picker.setEnabled(False)
        self.source_label = QtWidgets.QLabel(Dialog)
        self.source_label.setGeometry(QtCore.QRect(4, 240, 372, 20))
        self.source_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.source_label.setObjectName("source_label")
        self.source_label.setOpenExternalLinks(True)
        self.source_label.setParent(self)
        link_template = '<a href={0}>{1}</a>'
        self.source_label.setText(link_template.format(__source__,
                                                       'Source @ GitHub.com/digitalsleuth'))
        self.dauber_shape_label.setBuddy(self.dauber_shape)
        self.dauber_colour_label.setBuddy(self.dauber_colour)
        self.card_colour_label.setBuddy(self.card_colour)
        self.number_label.setBuddy(self.number)
        self.retranslate_ui(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        Dialog.setTabOrder(self.number, self.card_colour)
        Dialog.setTabOrder(self.card_colour, self.dauber_colour)
        Dialog.setTabOrder(self.dauber_colour, self.dauber_shape)
        Dialog.setTabOrder(self.dauber_shape, self.set_card_title)
        Dialog.setTabOrder(self.set_card_title, self.generate)
        Dialog.setTabOrder(self.generate, self.select_logo_button)
        Dialog.setTabOrder(self.select_logo_button, self.close)

    def retranslate_ui(self, Dialog):
        """Translate layout of the UI components"""
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", __description__))
        self.dauber_shape.setCurrentText(_translate("Dialog", "Circle"))
        self.dauber_shape.setItemText(0, _translate("Dialog", "Checkmark"))
        self.dauber_shape.setItemText(1, _translate("Dialog", "Circle"))
        self.dauber_shape.setItemText(2, _translate("Dialog", "Clover"))
        self.dauber_shape.setItemText(3, _translate("Dialog", "Heart"))
        self.dauber_shape.setItemText(4, _translate("Dialog", "Logo"))
        self.dauber_shape.setItemText(5, _translate("Dialog", "Maple-Leaf"))
        self.dauber_shape.setItemText(6, _translate("Dialog", "Moon"))
        self.dauber_shape.setItemText(7, _translate("Dialog", "Square"))
        self.dauber_shape.setItemText(8, _translate("Dialog", "Star"))
        self.dauber_shape.setItemText(9, _translate("Dialog", "Unicorn"))
        self.dauber_shape.setItemText(10, _translate("Dialog", "X-Mark"))
        self.dauber_shape_label.setToolTip(_translate("Dialog", "Choose the shape of the dauber"))
        self.dauber_shape_label.setText(_translate("Dialog", "Dauber Shape"))
        self.title_label.setText(_translate("Dialog", __description__))
        self.dauber_colour.setText(_translate("Dialog", "Red"))
        self.dauber_colour_label.setToolTip(_translate("Dialog", "Choose the colour of the dauber"))
        self.dauber_colour_label.setText(_translate("Dialog", "Dauber Colour"))
        self.card_colour.setText(_translate("Dialog", "Blue"))
        self.card_colour_label.setToolTip(_translate("Dialog", "Choose the colour of the card"))
        self.card_colour_label.setText(_translate("Dialog", "Card Colour"))
        self.generate.setText(_translate("Dialog", "Generate"))
        self.select_logo_button.setText(_translate("Dialog", "Select Logo"))
        self.close.setText(_translate("Dialog", "Close"))
        self.number_label.setToolTip(_translate("Dialog", "Choose the number of cards"))
        self.number_label.setText(_translate("Dialog", "# of Cards"))
        self.card_colour.setPlaceholderText(_translate("Dialog", "Blue"))
        self.dauber_colour.setPlaceholderText(_translate("Dialog", "Red"))
        self._menu_bar()

    def get_directory(self):
        """Get the output directory"""
        dialogBox = QtWidgets.QFileDialog()
        dialogBox.setFileMode(QtWidgets.QFileDialog.FileMode.Directory)
        dialogBox.setOption(QtWidgets.QFileDialog.Option.ShowDirsOnly)
        chosenPath = dialogBox.getExistingDirectory(self,
                                                    'Select the output location for your cards',
                                                    os.path.curdir)
        selected_dir = QtCore.QDir.toNativeSeparators(chosenPath)

        return selected_dir

    def select_logo(self):
        """Determine if the Logo option is selected and provide the result"""
        index = self.dauber_shape.currentIndex()
        if ((index == 4) and self.select_result.text() == ''):
            selected_file, _ = QtWidgets.QFileDialog.getOpenFileName(self,
                                                                     "Select the image to use",
                                                                     "",
                                                                     "Image Files (*.jpg *.png)")
            selected_file = QtCore.QDir.toNativeSeparators(selected_file)
            self.select_result.setText(selected_file)
        elif ((index == 4) and self.select_result.text() != ''):
            selected_file = self.select_result.text()
        else:
            selected_file = False

        return selected_file

    def allow_select(self):
        """Used to determine if the dauber selection option is checked"""
        return bool(self.allow_choices_action.isChecked())

    def on_select(self, index):
        """Modifies display of UI components depending on dropdown selection"""
        self.select_logo_button.setEnabled(index == 4)
        self.select_logo_button.setVisible(index == 4)
        self.select_result.setVisible(index == 4)
        self.dauber_colour.setEnabled(index in (1, 3, 5, 7))
        self.dauber_colour_picker.setEnabled(index in (1, 3, 5, 7))

    def dauber_colourpicker(self):
        """Colour picker for choosing the dauber colour"""
        dauberColour = QtWidgets.QColorDialog.getColor()
        colour_db = webcolors.CSS3_HEX_TO_NAMES
        if dauberColour.isValid():
            if dauberColour.name() in colour_db.keys():
                colour_name = (colour_db[dauberColour.name()]).capitalize()
            else:
                colour_name = (dauberColour.name()).upper()
            self.dauber_colour_picker.setStyleSheet(f'background-color: {colour_name};'
                                                    f' border: 1px solid black')
            self.dauber_colour.setText(colour_name)
        else:
            self.dauber_colour_picker.setStyleSheet(f'background-color: {self.dauber_colour.text()};'
                                                    f' border: 1px solid black')
            self.dauber_colour.setText(self.dauber_colour.text())

    def card_colourpicker(self):
        """Colour picker for choosing the card colour"""
        cardColour = QtWidgets.QColorDialog.getColor()
        colour_db = webcolors.CSS3_HEX_TO_NAMES
        if cardColour.isValid():
            if cardColour.name() in colour_db.keys():
                colour_name = (colour_db[cardColour.name()]).capitalize()
            else:
                colour_name = (cardColour.name()).upper()
            self.card_colour_picker.setStyleSheet(f'background-color: {colour_name};'
                                                  f' border: 1px solid black')
            self.card_colour.setText(colour_name)
        else:
            self.card_colour_picker.setStyleSheet(f'background-color: {self.card_colour.text()};'
                                                  f' border: 1px solid black')
            self.card_colour.setText(self.card_colour.text())

    def enter_title(self):
        """Enter the title for the card"""
        title_dialog = QtWidgets.QInputDialog()
        title_dialog.setWindowFlags(QtCore.Qt.WindowType.WindowSystemMenuHint |
                                    QtCore.Qt.WindowType.WindowTitleHint)
        title, clicked = title_dialog.getText(self,
                                              "Enter your title choice",
                                              "Title:",
                                              QtWidgets.QLineEdit.EchoMode.Normal,
                                              self.card_title.text())
        if clicked and title:
            self.card_title.setText(title)
        else:
            self.card_title.setText(self.card_title.text())

    def easy_mode(self):
        """That was easy.."""

        return bool(self.easy_mode_action.isChecked())

    def _menu_bar(self):
        """Add a menu bar"""
        self.menu_bar = self.menuBar()
        self.exit_action = QtGui.QAction("&Exit", self)
        self.exit_action.triggered.connect(QtWidgets.QApplication.instance().quit)
        self.file_menu = QtWidgets.QMenu("&File", self)
        self.file_menu.addAction(self.exit_action)
        self.menu_bar.addMenu(self.file_menu)
        self.options_menu = self.menu_bar.addMenu("&Options")
        self.easy_mode_action = QtGui.QAction("&Easy", self, checkable=True)
        self.easy_mode_action.triggered.connect(self.easy_mode)
        self.options_menu.addAction(self.easy_mode_action)
        self.allow_choices_action = QtGui.QAction("&Allow Dauber Selection", self, checkable=True)
        self.allow_choices_action.triggered.connect(self.allow_select)
        self.options_menu.addAction(self.allow_choices_action)
        self.help_menu = self.menu_bar.addMenu("&Help")
        self.help_content_action = QtGui.QAction("&Usage", self)
        self.help_content_action.triggered.connect(self._help_menu)
        self.about_action = QtGui.QAction("&About", self)
        self.about_action.triggered.connect(self._about)
        self.help_menu.addAction(self.help_content_action)
        self.help_menu.addAction(self.about_action)

    def _help_menu(self):
        """Add a help menu to the menu bar"""
        self.help_box = QtWidgets.QDialog(None, QtCore.Qt.WindowType.WindowCloseButtonHint)
        self.help_box.setWindowTitle("Help")
        self.help_box.setFixedSize(610, 730)
        self.help_label = QtWidgets.QLabel(self.help_box)
        help_font = QtGui.QFont()
        help_font.setPointSize(10)
        help_font.setFamily("Arial")
        help_font.StyleHint("SansSerif")
        self.help_label.move(10, 10)
        self.help_label.setFont(help_font)
        text = (f"# of Cards:\tChoose how many cards you would like to generate - must be a "
                f"number\n\n"
                f"Card Colour:\tEither click the colour box to open a Colour Picker dialog box\n"
                f"\t\tor type the colour into the text box. Colour names are drawn from:\n"
                f"\t\t{__colour_groups__}\n\n"
                f"Dauber Colour:\tAs with Card Colour, click the colour box for the"
                f" Colour Picker\n"
                f"\t\tor type the colour into the text box. Colour names are drawn from:\n"
                f"\t\t{__colour_groups__}\n\n"
                f"Dauber Shape:\tClick the dropdown box to choose the shape of the dauber to be"
                f" used\n"
                f"\t\tThe 'Logo' option will present the option to choose a custom file or logo to"
                f" use\n"
                f"\t\tas a dauber. This file will be resized to 48x48px, so ensure the image you"
                f" choose\n"
                f"\t\tis of a good quality to start.\n"
                f"\t\tThe Dauber Shape option works in conjunction with 'Allow Dauber Choices'."
                f"\n\n"
                f"Card Title:\tThis allows you to put a title or text banner at the top of the "
                f"card, which\n\t\tappears under the card number - eg. 'BINGO BONANZA 2023!'\n\n"
                f"Generate:\tWill generate the bingo cards with the selected options. "
                f"If some options\n"
                f"\t\tare left blank, defaults will be selected (blue card, red circle dauber).\n"
                f"\t\tWill then pop up a directory selection box to choose where to save the files."
                f"\n\n"
                f"Close:\t\tSelf-explanatory, will close the application.\n\n"
                f"Select Logo:\tThis option only appears when 'Logo' is chosen from the"
                f" 'Dauber Shape'\n"
                f"\t\tdrop-down box. Opens a File Chooser dialog to select either a"
                f" JPG or PNG file.\n"
                f"\t\tThe file chosen will show in the box below to confirm the choice.\n\n"
                f"Options Menu:\tEasy means when a number is selected, it is selected on all spots "
                f"that contain\n"
                f"\t\tthe number which was clicked. If B4 is called, when the player daubs it on\n"
                f"\t\tone card, it is automatically 'daubbed' on all cards on the sheet.\n"
                f"\t\tIf this option is not selected, play is as per normal.\n"
                f"\t\tNOTE: This option is not presented to the player, and is only available "
                f"within\n\t\tthe application.\n\n"
                f"\t\tAllow Dauber Choices presents the player the option to choose their own "
                f"dauber\n\t\tfrom a drop-down box at the top of the card. "
                f"Option is not available if not selected.\n\n"
                f"The final output of the application will be a combination of HTML files, "
                f"PDF files and a single Excel\n"
                f"spreadsheet. The HTML and PDF files will be named for the card number and "
                f"colour\n(ie. 1-BLUE.html, 1-BLUE.pdf). Both the HTML and PDF file should be given"
                f" to the player so they have\nan option to print or click.\n\n"
                f"The Excel file is named for the card colour, and allows you to enter the numbers "
                f"called into the CALL\n"
                f"sheet, and when Bingo is called, you can click on the sheet with the card number "
                f"on it and the called\nnumbers will be automatically highlighted. "
                f"This enables easy confirmation of a successful BINGO!")
        self.help_label.setText(text)
        self.help_label.adjustSize()

        self.help_box.exec()

    def _about(self):
        """Add an About menu to the menu bar"""
        self.about_box = QtWidgets.QMessageBox()
        self.about_box.setIcon(QtWidgets.QMessageBox.Icon.Information)
        self.about_box.setWindowTitle(f"About {__description__}")
        self.about_box.setText(f"{__description__}\t\t\t\n"
                               f"Last Updated: {__date__}")
        link_template = '<a href={0}>{1}</a>'
        self.about_box.setInformativeText(link_template.format(__source__, __source__))
        self.about_box.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
        self.about_box.exec()

def gui_everything(number, card_colour, dauber_colour, dauber_shape,
                   output, logo, allow_select, easy, title):
    """Takes all input from the GUI and passes it to the various functions"""
    args = {'num': number,
            'pdf': True,
            'card_colour': card_colour,
            'dauber_colour': dauber_colour,
            'dauber_shape': dauber_shape,
            'logo': logo,
            'allow_select': allow_select,
            'base_colour': card_colour,
            'output': output,
            'easy': easy,
            'title': title,
            'excel': f'{str((card_colour).strip("#"))}-cards.xlsx',
            'everything': True}
    if not output:
        return
    create_card(args)
    grab_numbers(args)
    generate_excel(args['num'], args['card_colour'], args['excel'], args['output'])
    if number == 1:
        amt = 'card'
    else:
        amt = 'cards'
    if dauber_shape.lower() not in ('circle', 'square', 'maple-leaf', 'heart'):
        dauber_colour = ''
    else:
        dauber_colour = f'{dauber_colour.lower()} '
    results_msgbox = QtWidgets.QMessageBox()
    results_msgbox.setIcon(QtWidgets.QMessageBox.Icon.Information)
    results_msgbox.setWindowTitle("Finished")
    results_msgbox.setText(f"All files created in {output}\n\n{str(number)}"
                           f" {card_colour.upper().strip('#')} {amt} created with a "
                           f"{dauber_colour.upper().strip('#')}{dauber_shape.upper()} dauber.\n"
                           f"Excel file named {(args['excel'].upper())} created for tracking "
                           f"called numbers.\n\n"
                           f"You may now close the Bingo Card Generator, or generate more cards.")
    results_msgbox.setStandardButtons(QtWidgets.QMessageBox.StandardButton.Ok)
    results_msgbox.exec()

def create_card(arguments):
    """Creates the HTML version of the card"""
    card_colour = arguments['card_colour'].lower()
    dauber_colour = arguments['dauber_colour'].lower()
    dauber_shape = arguments['dauber_shape'].lower()
    output_path = f"{arguments['output']}{os.sep}"
    cards_per_sheet = 6
    if not os.path.exists(output_path):
        os.mkdir(output_path)
    total = 1
    if not arguments['title']:
        card_title = ""
    else:
        card_title = arguments['title']
    open_head = '<!DOCTYPE html>\n<html lang="en">\n<head>\n<meta charset="UTF-8">'
    open_style = "<style>"
    if not arguments['logo'] or arguments['logo'] == '':
        selected_logo = 'url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACgAAAAyCAYAAAAus5mQAAAABmJLR0QAAABxAK/qGQdhAAAACXBIWXMAAA7DAAAOwwHHb6hkAAAR4ElEQVRYw81ZaZgU1bl+q7qqeqnee3p69n0GZoMBhh0ElEUGUUEMKBhQk5gYDSEGvUluNpN4Ta5LNKKicQkkcQkgi8ggkgEHGGZhmH3fp6ent+l9ra4lP7iPXh/jFaLeJ+d31Tnveb/3/c53zkfgSxxN9ecJU6xZMz4RCF1350/EL2NO8ssCNzk8RrCJ8WVC0H0qLT/v5mNvvUv8WwH0hhMAFygNh8NJBBdOo+Vy/FsA7O7uIAGgpCxfCoQ9o35BnjnS3+G68dZVEgBcvnCG6u9q+pfZ/Jd+rL94QU6I0TyDTn99wOeamZyec2JybLQpNNawNzHesZbXpdVYKrds1bKy+bFwYJ2KZds5gawW4+Jo2dz58a8U4PvVhxiLOW0ZN1n/uIJ39xMy2hPjyPlhOn1ccvXO1FGhnCC0g6Iur0lDuPMYmrmk1DCacJwpoIzlj3ntk9VLqjZyX1mIjVqzWfL2/EwTHnh90CZtLVv/s/tJkVivnzrbKBet9oRAgvB2ezTO2g4jI6sqr/rBd9Rs+XZWkXgtMt70kDE9K/krZbCx9uRdlKvuoZF+dhVz4E1KlpO9XeaYnO81GP5GVqYnp0/PfDrgFXcGTzRNqR32rwX1xg4MDTxveGi3aNSNHTeUbfx5bsncY1e7HnWtAEOhIKXhEnSkpvplw8joWjgniAQfp9lJ/63j1gyk/OhejO599RntZSuQZ4FuoP8mIso/Mnns8Ie628uYUDROfqUuNhhTjkXkGT63MHrThNfL+LK1dOyuRXCUmBGKEIg2HYc/LCA0PRWK+9fARodJazQuS7C+5aSx3MMoNeeuZb1rZtCckuKxXX73rdL5mYtGdRZkr8gGJxFg75gHo3US2oJ5mF7VBlNRCSLuIRT9eDtaD9XDoo4TCb/toJFa7P1KGUzPLhAF7bSTptK5cUMSh7aDDYja7eg424PxThe4wCCGm+2oP1SDQJxGw5tnYEglkDlrVsQnGY+aiwqv6QiUXSvA2rMjRCjiCMYHmjKEsHM2DEZYspRIm1EINoUBa9ZBX5SNoqUVoKQIRIaBURWEkiHfdKdv3L9u3Qbp8DtvfvkMvv3H32bVfnCSNSvOVpRYhp4iTZkRNRuSMrNVsHljEPQ5SJm9DK/U0tBY0hElVXD4YshKI6FK1iKaXslnTbzybInr5LymM8fY9//2bO6XxmDz2ddTM9W21zzv/HG3IU9ZqONdO0ZPt3g9QZbJmrdMn1qyAEGOQU7FapiSLMiYVglGZYZCaYA+KQvuIV99RkUyk6YXt9h7BvTZpfi13GNdtnX1ypOvvHsu9IVNYvfSsUW5isJYBpuTrMW0vneaJeWMm2LFK9YmTU1OQEURMMqG4Dj/NpISHKJhNQi5DnplEiRTOfK2ryzydR85mVmkQbIxcptOIRKT7c1RoWR79Aslap+7l3BNjOsGWxt0BSn+TZqE8wmeA0JuNxL5y0XXpS5SmxiFKxTH2LAV9jCBMCioCAHJSh7ZhXlIkgkQsioRtw2IhbkCSbJ6iAKBzmb/D3M23XsgmtBOhSOa8KLlc6VrZrC7o8ucqR15QzvygTmgWPYIm5J6ynGhfgVdMo+S2YbJtuZzOGDl0S5jECWU/0vQFKSAHJTdjlIhgc15p1B14wLSEY1DzXtELkIf1s5dMqZNdFU7jlRPkDM33QHAddUAu84ckQVttS+ZZT28muJLFVw4mWb5m0O8MaKeszxmbTyn/s2JfpyjFCApGfQAlmsJ9PsFDBHkR6ERCAJtFIOWMQHvvFCDHy1NR/6COULAbU/EO9vW6fLzC+mII9WS4n2q/fgvAwEu5YHFG+6TPtfFjc1dhnBjbZVadH2LlissCrOOSE8lvk3J+Ft6ak+qv1E9gguU4qOfi1gSv7y7HPsfnIU1hk/u+Xo9gxkKCk0yBt+qtePSqRo6KT95M22t2R4P+mRJOWadgQls41rOrYK2kL0qFx8+eTqyperWulDzhzfrMlNYcGH4pWSMN57BN2qccJKf3NdEQkJjiwvrFxXh+socnL80ATd/JR/fVmTE3asL0XbZiTGSQI09gWUqH8yzlgKecSh1atib+8dciZQty+/ZPXJVAC+feW3hlNsXMJUV2PyjVrdX0E5TE1HqwbdaMAwZ1IIIPyeAJwmQBAESgEOQUHtpAvEpDhsWF+J8zySiElDOMth080IcPNWOSQngCAItw0FsnW+GmzP6pWjszzFzwXEyu2hsy/LZBfvfPT/6f4bYafdQJk342Wyhp84+GApTeQWsOSeTeefI+2gBjQVqCr+/pQCDv1qBht2r8KfNy/DrGyqxc04RytPy0RM04NQlB+6dmQUCgDMsIBInYCM+Thi9BIWDJ+uRlKZlB1u8cjYpky9O52vo7r8//srujfRnmmSi95ys5YNn55cXq4dTytIq6XHnM1qVWhmyBbB3jIJEAwuyVFBrWNT3BjHpdOL1v3dj7ZL1uPveXUhJNsPz4TmoFyzA0FA/xv0v4sRQD5panYhKBICP9f9iP4mNiQClgusuub9jYzBAqmQaizW7pHShe/Thc0nZC8RPAbS1nFtaqrO/oRQNRiZVj7jToUyEougf7MEkTYMA8PuuANDVDgBgBAkPr9+K733nPsgZBn2//R3sx6uRt/N+zNm0CZa0DIzufgiDVi9C4hVNlsopyAigJQb0DLmQMasMag3Dulp7UbKyYlMsHl462n5xJYCOT2mw1JIqKHs7p/nbmnMJkpAnQHNgtfGG9j76tEP4hBQKlCpsLinB93c9jGhnJxzvVSM6PIapoyegmz8PgcEhJGVnoLCwGLXnT6PFG0AhBbz63UoEA3E0OsIo11Aom1kixn0BgaEI0t4+5He0244OhVIPHTp9NvopBr+/Z/94c+3fd9jbTq6RUeGMCKGcma3Gjq7RAIAr91wJwLr0JKxfUIEgsqDRaDDcN4jhBx8GSZEQSBIT//U7JDx+qC/WYOaMMrz1lgYpJIFnd1RAQVM40+EGAAyMOUEp5WRjbXTvtKKM7jFNcU/66qq6rQuXhT7TJGTf0Ttmz1Y9py+YISucprxFrlJS9gT9UfLdWZGFn359Id54/wL0JhMACWk3rYXhgfsQ4QWESAlRXxAZTz8OQ2kxFAo5rL44qrLUsJhVeOZgB1r/J9zeKA+GFsAOntpqLMjiFi5Uvp7oOPa1zzTJ3/Y8LgsPvL84GKLSkpcv/pUgQRkTGGgUDBDmsd7M4of3rERjmxUd4Rj4eOzKJGoWlm/uQEAtBzdug0ohR8rGDSAZOWLhCJiYB0sXpuJ7zzfhnD/+ESsqWoZEnINCq9VoFaGnbGeblNGJ4Lw3X35635Zv7uI/pcFdP3kc2aXKmH/cOoefHMlQZ6SS/ikfXCEJH457cM+yQjAUg2+9dApuQUB6JIilS1bjiaf/gMdefRFOjGAQQbQyPPbsfwcWnQkyXsJo+zG80GhHZ+yTOr6lJANFFjlSMlMIf/8YzZNUi3Jh1T5TwY2de1/c8+kQk6HmhYRv9K20aeYSmubhaW1DyOEO52ZZAADvXhjEt/9wHDaeB00ABrUHbx8+gjhIDPN+LMg0on7Mj0KNiM6AGx6vHy/vexkvjUQwIX2ycJIAFOTkQgyHOD4UhkonQ0ZJyiyzaN0jeZvL/2mIo2Rhfe+pI3sZ3p6uK8gokBRpf3X5px4uzk1h5aKEWv/HrxZsQoBRLcNP9j+H327/LhpeOABGimH17VFIFIvVt0fw0l9fw4nORsw1GUASIiAJGAnEYBMEsLyIGSW56Dj+3nBmXtY+RTSy3tFja/BzsUDBfzzWAWz77Hrw2KHnFEaGX0uCT2SZvAfjUx7m6ROD+GOX9aNvrmMpPPfDNdj8aDUckoA7S8sgKi3QaTSIRwJIxKeQpJOjtFCPipk5gMjD6XTjx8+fxhlvGA/OnoH7VqRh/OC+UHzR179OG7TJAtg/r9q0O/y55RZr69ueW6J+UmlS+uQKFdHV3oO7bijG4V4b3MIVBxZn6mFJZqFXyNAdFdA2MYwNS4w4UluDWr8fIUkCmxCQV0thul4JllWhftiFYYpEGkjcuWomBo++An1OrqJ8jvbVqMsn7+iZcgM4+LnFwqa5M9ICZ07cmPAFU2mNRqY0qBAatUszp+cQRzsnQBAAFYhhxewMfHBxCCOciNFYAtkKCr+4/wYUqXU40zMGgSLhkYC+KI/OQBQ+kgAD4LkdG6B19gpJ6XrSVFRAWmvqFJPN/cOSpeIv+987O/65AF86fbQ/Lo/IWTV3XdwxSUqhmNcaUncXWZjUHD2Lk4Mu2EXAFA6i3eqHXbyilGaHF/myGK5fMg2Ls82obRtHRPr4/JVJwJO3VaFCw2FgPFhn1pImUogx8uxUjqlY8eiiLT89eFXXTk+fnUHQvcE7MElNNPRxYZ7+S/7c6ecotQyrMlV4edNiJBEEft3kwGWe+IQz/7O6F/f9/AAmrU4k0R9PnyWj8cr2jVicLGH04AuQ2z31Xj8O9L7XwDtaR2gx4FkzaXPIrqrkF2Uc19/CPEW8WP8XQgQhzV52ndHvmcjPS5VGglGimA9IB3euJ/50uhVvt48hAOkjt0UJAhc5HnXvd155ywGwtSwLd61bDr6nURz/4AyZXFQiKpbMLRm80JMpHpokXbRTSuzIf7kylxev6lZXP9hJWPc+fAF1bQs0i4uRkpOM1AId5GoGoUgM1uEIF/Yx1XoLURVRqqjL3RM42TiM0WAMkiSBgIRsjRxrZmejojgNLJfAhDXWnJTCFphNpNaglyMWlDDY2IGYxCB4sQ9i5ex9m37z9varAnh4zy8KLBMnOmVmI5NRlOofrBlpVFNhC6mKl18OJqPdwYEkJAASeF4ECBIkSYKmCEiiCAhxcDwgCABBUqDkckBIgJdIXJcnx9xSI0bqhlslubGz6Prsm4e6J9R8f0skULkrZ/09u1yfq0HW1nm7QuZjFHoGQU61J2nbj+/U37SSMs2cA78/jtYhF3xxAVNREVNREXZ/DDZPCKSKgTsuIkgoYQ+LcMQBY6YBx850ISnLhGf+egnxqAiRpJB+2yrJN/+B77hC5CEtK0Kjk6tkPbV3XJVJaAJzFAYT9Col4hK5sOfyxaJg75C1r6bjrEgxAUEC5AoKpiQNtGoGcqUSDi+HYISDSBAoyjODoChotSooaBm23T4HoRAHEAR4iK6BDzvPujvGxslwyyxGTswypyVDk5mNeGBs+lWZJEqpDqjMlg20UkVmWbQrTGn8G0PW/LvNd2xulO17ts81FdYmspIwPGqDnBYRCAuYmgqjf0CC0xvGQP8kaBmJQJRH3SUJkiRCp1YAtAwkLfOIi3avoZj2tQtSuQMq2pzEeaYQ0WgFPr3gXaDp8zUoSRLR9Op3X8otYO9VmEzEwLgHIh9PaHSG4z117Us5QjmV4OKyOJewyMSoWuAFUZWcMU5CFMALgChAEEUqASITiSghJXi3giH8ECVSJonK9NklzRTB31hWkk1ClBANeMWBsdgTczY++chVMUgQhPTaq/se0GJqxCDhkVCE0+h1Stpud9867/pybsIpXoiL8rcnbYFCbWg40945MNlT/NRv1PnlAqG+su1wZwezxvLBL2OOSX1ze+DivMoUMUmv2eZw2ZboTIaqzo5BGI1emIxa35CD/Jk97QfPA09e+yt/Q/UTSzJSqWfiEd/sS00DULAqlM+YBokPwWbzRsyW9Eaf196qYg3dNpvfFgpNRLWsikrPzMzmYpEZsZhQQpPhhV5/nMkvzEVD8yjSUnRQMASUjOJDm4f+wcrNP7/0hdoQredPKTlf3T06JbfTZvcUlJTlE8PjLtAkIAgCFCoKbZeHcNuGJdLImBNO9xQIkScoRgmNRolENAaRoKBUMFCzCpFLoMvplf23fYp6Y+PdjyS+tD7J0T89x2aZfav0WnwjkkjMV8hJkyQKRN/wFLiEhHWrZ2Fg0IZI0A+GphGMJTAtx4ThEaeYmaZ3TTgjdWFO/8KIQ3V227d3XVVL7F9u8v35D79KKcgIV+bkphUE/L45jIzPs6QYyFiMh5jgEAqGBZs9OJhbmNXQ2jw27AoZGrftfNT1/9JM/Gdjy3sSOW8AKJ4JHD4CWO4BHi0jvnBT+x8+c0PISk6iJwAAAABJRU5ErkJggg==);'
    else:
        selected_logo = f"url({convert_logo(arguments['logo'])});"
    page_css = '''
:root {
  --logo: ''' + selected_logo + '''
  --github: url(data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAACgAAAAoCAYAAACM/rhtAAABhGlDQ1BJQ0MgcHJvZmlsZQAAKJF9kT1Iw0AcxV9TpSIVETuIOASsThbEijhKFYtgobQVWnUwufQLmjQkKS6OgmvBwY/FqoOLs64OroIg+AHi6uKk6CIl/i8ptIjx4Lgf7+497t4BQqPCVLNrElA1y0jFY2I2tyoGXiHAjwGMIioxU0+kFzPwHF/38PH1LsKzvM/9OfqUvMkAn0g8x3TDIt4gntm0dM77xCFWkhTic+IJgy5I/Mh12eU3zkWHBZ4ZMjKpeeIQsVjsYLmDWclQiaeJw4qqUb6QdVnhvMVZrdRY6578hcG8tpLmOs0RxLGEBJIQIaOGMiqwEKFVI8VEivZjHv5hx58kl0yuMhg5FlCFCsnxg//B727NQnTKTQrGgO4X2/4YAwK7QLNu29/Htt08AfzPwJXW9lcbwOwn6fW2Fj4C+reBi+u2Ju8BlzvA0JMuGZIj+WkKhQLwfkbflAMGb4HeNbe31j5OH4AMdbV8AxwcAuNFyl73eHdPZ2//nmn19wOfLXK54rfCmAAAAAlwSFlzAAA3XQAAN10BGYBGXQAAA8hJREFUWMPVmE1oVUcUx3/vkkXc1GxMFsaNtgubB360wY9CqA3YLEpAUHGhVyJXL6KC2grVgJ/oRrC1BMLIqDAtKBJQESRGJHRjlGK1kOjC6MKsEjfqps8u8tycF4ebd+fe93HVDrzFuzN37m9mzvnPOQc+8Zar9kWjVROwEugAPgdagc+k+w0wATwG7gP3/CB8lTmg0SoHrAc2A2uBxpSvFoAh4A9gwA/CYl0BLbDDQL7GUxsFjqcFzaWAWwCcA7rqbF6DwA4/CCeqBjRarQEuA80Z+cAUsMkPwuGKAY1WXcDVCuys2lYA1vlBOJga8APCJULmysDlgb8+IJwN2e4H4aj90IvANQKXysC9Ar4GtgO3gGKVEHeB3TLXeKSvEbgkDDOtITJoX4yM3PaD8AHwANCyy78Ba4BJ4AnwUn5vRbTnAYvFwR4CP9nOYLQaEoG3W16k7NAsQJGTwzErH7P/+EE4arTqBBYBz+L0TPSzFXjpB2HBNae9SUar/pL8NER2L7XdCdR4ijFxOvdfzPNGYdk/Y4Ny7j2Ob7Vl4BRfOvp6SrZYcpK1QJPjhWIGgK45m4RpBrA7Qe3DDAB7gaeO/m4bsMMx8NdqQ6UE+ywAJxxDOgA8o5Un3hh3DAMZivMNx1EvMlrN8USrPAfgs6zo5GTi5veAhV7CHE/8IJzO+Ip76+pMAlwsJvDRWtLHc1aeUfcmN838WgGXZrhBC4C5ruP35KJ32dn3GQL+4Aiap4HnnjjBI8ck24xWczM63p2OIY/8IJwuHfGIY2AzcCqD3fs5IUMcsW0wSYx3Gq0OyqrrsXu7gJMJwwZswD+BF2VsoGg5yyngptHqqxrA8kara0BfQkb5QpjeDzJaHQGOWjdIh1zmp4EtkQn+lgkeAiN+EI7HALUCq4EVwLfAspTFgqN+EB6LBqxngb0S6uSA25I/bJV+G3K5/ApAe4JMnUuQknLR09lZOij3Ym8ksu2TYHWPlCxm2Uk0C4vctRPAhQot4ZAdPUWFuh8YjkCe9oPwNdAJXAH+FfscA86n+OC9CuCGowsqlxe3AP8ALZY9tktWR+luThtEGK02yMKS2iSwxA/CyTSVhTxwx6rJTElOfMPO4IxWuaQKVUrAKaCznLk0xNhOKa0sQTYD14FJo9VzCZHmidfXGtDGwjmDBXnhm4hztACrRDLaqKFCa9UKv3M5mpcQ8Y6LjJwRSalnRveL2PaYa2BDyuTmR6PVRak8rK9h54pS2zng2rWKACNHvtFo9YUk+a2l6yihDQG/S93mYlqw/017B09eQK3tE7jIAAAAAElFTkSuQmCC);
}
.clear {
  width: 100%;
  clear: both;
}
.card {
  margin-top: 25px;
  float: left;
  width: 400px;
  height: auto;
  background-color: ''' + card_colour + ''';
  border-radius: 5px;
  padding: 10px;
}
.clear-card {
  position: absolute;
}
.card-title {
  position: relative;
  justify-items: center;
  text-align: center;
  font-family: "Roboto Condensed", sans-serif;
  font-weight: bold; color: ''' + card_colour + ''';
  font-size: 40px;
  margin-bottom: 10px;
}
.card-number {
  position: relative;
  justify-items: center;
  text-align: center;
  font-family: "Roboto Condensed", sans-serif;
  font-weight: bold;
  color: ''' + card_colour + ''';
}
.grid-container {
  display: grid;
  grid-template-columns: auto auto auto;
  grid-gap: 5px;
  grid-template-rows: auto;
  grid-row-gap: 0px;
  justify-items: center;
  align-items: center;
}
.grid-child {
  float: left;
  justify-items: center;
  align-items: center;
  padding: 5px;
}
.center {
  display: flex;
  justify-content: center;
  position: absolute;
}
.headers > div {
  float: left;
  width: 80px;
  text-align: center;
}
.headers > div span {
  font-size: 30px;
  color: #fff;
  font-weight: bold;
  font-family: Arial, Helvetica, sans-serif;
}
.column {
  float: left;
  width: 80px;
  text-align: center;
}
.number {
  padding: 20px 0;
  border: 2px solid ''' + card_colour + ''';
  background-color: #fff;
  font-family: "Roboto Condensed", sans-serif;
  font-weight: normal;
  height: 16px;
}
.number:hover {
  opacity: 0.8;
}
.number span {
  color: #000;
  font-size: 20px;
}
.number span:hover {
  text-shadow: 0 0 5px rgba(0,0,0,0.5);
}
.button {
  background-color: ''' + card_colour + ''';
  border: none;
  color: white;
  padding: 8px 16px;
  border-radius: 8px;
  text-align: center;
  text-decoration: none;
  display: inline-block;
  font-size: 16px;
  margin: 4px 2px;
  cursor: pointer;
  transition-duration: 0.4s;
  width: 100%;
}
.button-clear {
  background-color: white;
  color: black;
  border: 2px solid ''' + card_colour + '''
}
.button-clear:hover {
  background-color: ''' + card_colour + ''';
  color: white;
}
select {
  display: block;
  margin: 0 auto;
  text-align: center;
}
.circle:after {
  display: flex;
  content: '';
  margin-top: -25px;
  width: 35px;
  height: 35px;
  background: ''' + dauber_colour + ''';
  border-radius: 50%;
  position: relative;
  top: 10%;
  left: 27%;
  box-shadow: 0 2px 4px darkslategray, inset 0.1em 0.1em 0.1em 0 rgba(255,255,255,0.5), inset -0.1em -0.1em 0.1em 0 rgba(0,0,0,0.5);
  -ms-transform: translateY(-50%);
  transform: translateY(-30%);
}
.square:after {
  display: flex;
  content: ''; margin-top: -25px;
  width: 35px;
  height: 35px;
  background: ''' + dauber_colour + ''';
  border-radius: 10%;
  position: relative;
  top: 36%;
  left: 27%;
  box-shadow: 0 2px 4px darkslategray, inset 0.1em 0.1em 0.1em 0 rgba(255,255,255,0.5), inset -0.1em -0.1em 0.1em 0 rgba(0,0,0,0.5);
  -ms-transform: translateY(-50%);
  transform: translateY(-40%);
}
.maple-leaf {
  display: flex;
  align-items: center;
  justify-content: center;
  position: relative;
}
.maple-leaf:after {
  position: absolute;
  content: '';
  width: 40px;
  height: 40px;
  background: ''' + dauber_colour + ''';
  -ms-transform: translateY(-20%);
  transform: translateY(-1%);
  clip-path: polygon(47% 100%, 48% 70%, 25% 73%, 28% 65%, 7% 47%, 11% 44%, 8% 30%, 20% 32%,
                     23% 27%, 35% 40%, 32% 13%, 39% 16%, 50% 0, 61% 16%, 68% 13%, 65% 40%, 77% 27%,
                     80% 32%, 92% 30%, 89% 44%, 93% 47%, 72% 65%, 75% 73%, 52% 70%, 53% 100%);
}
.heart:after {
  display: flex;
  align-items: center;
  top: -45px;
  margin-left: 18%;
  position: relative;
  content: "\\2764";
  font-size: 45px;
  color: ''' + dauber_colour + ''';
}
.star:after {
  display: flex;
  align-items: center;
  top: -45px;
  margin-left: 15%;
  position: relative;
  content: "\\2B50";
  font-size: 40px;
}
.moon:after {
  display: flex;
  align-items: center;
  top: -45px;
  margin-left: 13%;
  position: relative;
  content: "\\01F319";
  font-size: 40px;
}
.unicorn:after {
  display: flex;
  align-items: center;
  top: -45px;
  margin-left: 10px;
  position: relative;
  content: "\\01F984";
  font-size: 40px;
}
.clover:after {
  display: flex;
  align-items: center;
  top: -45px;
  margin-left: 10px;
  position: relative;
  content: "\\01F340";
  font-size: 40px;
}
.logo:after {
  display: flex;
  align-items: center;
  top: -40px;
  margin-left: 23%;
  position: relative;
  content: var(--logo);
  font-size: 40px;
}
.x-mark:after {
  display: flex;
  content: "\\274C";
  margin-top: -20px;
  margin-left: 3px;
  position: relative;
  font-size: 50px;
  -ms-transform: translateY(-50%);
  transform: translateY(-50%);
}
.checkmark:after {
  display: flex;
  content: "\\2705";
  color: green;
  margin-top: -19px;
  margin-left: 4px;
  position: relative;
  font-size: 48px;
  font-weight: bold;
  -ms-transform: translateY(-50%);
  transform: translateY(-50%);
}
.footer {
  opacity: 0.7;
  display: flex;
  justify-content: center;
  content: var(--github);
  margin-left: 49.25%;
  height: 40px;
  width: 40px;
}
.footer:hover {
  opacity: 1.0;
}
'''
    close_style = "\n</style>\n"
    close_head = "</head>\n"
    open_body = "<body>\n"
    if arguments['allow_select']:
        select_box = '''
<select id="dauber" name="dauber" class="selectpicker">
  <option selected disabled>Select a dauber</option>
  <option value="checkmark">&#9989; Checkmark</option>
  <option value="circle">&#128308; Circle</option>
  <option value="clover">&#127808; Clover</option>
  <option value="heart">&#10084; Heart</option>
  <option value="maple-leaf">&#127809; Maple Leaf</option>
  <option value="moon">&#127769; Moon</option>
  <option value="square">&#128998; Square</option>
  <option value="star">&#11088; Star</option>
  <option value="unicorn">&#129412; Unicorn</option>
  <option value="x-mark">&#10060; X Mark</option>
</select>
'''
    else:
        select_box = '\n'
    body_title = f'<div class="card-title">{card_title}</div>'
    footer = '<a href="https://github.com/digitalsleuth/bingo-card-generator" class="footer"></a>'
    script = '''
<script src='https://cdnjs.cloudflare.com/ajax/libs/jquery/3.6.3/jquery.min.js'></script>
<script id='rendered-js' >
'''
    if arguments['easy']:
        emode = '\nvar easy = true;\n'
    else:
        emode = '\nvar easy = false;\n'

    js_array = "$i = ["
    for card_num in range(1, (cards_per_sheet + 1)):
        js_array = js_array + str(f"$card{card_num}, ")
    js_array = f'\n{js_array.rstrip(", ")}]\n'

    js1 = '''
$(document).ready(function() {
  var idArray = [];
  for ($card = 1; $card <= ''' + str(cards_per_sheet) + '''; $card++) {
    for ($x = 0; $x <= 24; $x++) {
	var numberId = ("card" + $card + "-c" + ($x+1));
	idArray.push(numberId);
    $("<span>" + $i[($card - 1)][$x] + "</span>").appendTo('#card' + $card + '-c' + ($x + 1));
      }
  }

function toggle(id) {
  var element = document.getElementById(id);
  var dauberChoice = $('select[name=dauber]').val();
  $('select[name=dauber]').prop('disabled', true);
  if (dauberChoice == null)
    { dauberChoice = "''' + dauber_shape + '''" };
  element.classList.toggle(dauberChoice);
};

function easy_mode(clickedNumber) {
  idArray.forEach(function(cardId) {
    var currentNumber = ((document.getElementById(cardId)).innerHTML).replace( /(<([^>]+)>)/ig, '');
    if (currentNumber == clickedNumber)
	{ toggle(cardId) };
  });
};
'''
    js2 = '''
  $('.number').click(function() {
    if (easy == true) {
  var callNumber = (this.innerHTML).replace( /(<([^>]+)>)/ig, '');
  easy_mode(callNumber);
  } else { toggle(this.id) };

  $('#clear-card').click(function() {
    location.reload();
    });
  });
});
</script>

'''
    header = ['B', 'I', 'N', 'G', 'O']
    columns = {1: [1, 6, 11, 16, 21],
               2: [2, 7, 12, 17, 22],
               3: [3, 8, 13, 18, 23],
               4: [4, 9, 14, 19, 24],
               5: [5, 10, 15, 20, 25]
              }
    if arguments['pdf']:
        print("Generating PDF's, please wait")
    while total <= int(arguments['num']):
        filename = f'{output_path}{str(total)}-{(card_colour.upper().strip("#"))}.html'
        title = f"<title>CARD {str(total)} </title>\n"
        count = 1
        card_clear = ('<div class="card-number" id="clear-card"><button class="button button-clear">CARD ' +
                      str(total) +
                      ' - CLICK HERE TO CLEAR CARD</button></div>\n')
        free_space = (f'$(\".col-13\").html(\'<span style=\"color: {card_colour};'
                      f' font-weight:bold\">FREE</span>\');')
        html = open(filename, 'w')
        html.write(open_head)
        html.write(title)
        html.write(open_style)
        html.write(page_css)
        html.write(close_style)
        html.write(close_head)
        html.write(open_body)
        html.write(card_clear)
        html.write(select_box)
        html.write(body_title)
        html.write('<div class="grid-container 1">\n')
        if cards_per_sheet > 3:
            second_grid = int(-(-cards_per_sheet // 2))
        else:
            second_grid = ''
        for card in range(1, (cards_per_sheet +1)):
            html.write(f'<div class="grid-child {card}">\n')
            html.write('<div class="clear"></div>\n')
            html.write(f'<div class="card {card}">\n')
            html.write('  <div class="headers">\n')
            for letter in header:
                html.write(f'    <div><span>{letter}</span></div>\n')
            html.write('  </div>\n')
            for col, _ in columns.items():
                html.write(f'  <div class="column {col}">\n')
                for colnumber in columns[col]:
                    html.write(f'    <div class="number col-{colnumber}"'
                               f' id="card{card}-c{colnumber}"></div>\n')
                html.write('  </div>\n')
            html.write('</div>\n</div>\n')
            if card == second_grid:
                html.write('</div>\n<div class="grid-container 2">\n')
        html.write('</div>\n')
        html.write(footer)
        html.write(script)
        while count <= cards_per_sheet:
            nums = str(generate_numbers())
            html.write(f'$card{str(count)} = {nums};\n')
            count += 1
        html.write(emode)
        html.write(js_array)
        html.write(js1)
        html.write(free_space)
        html.write(js2)
        html.write("</body></html>")
        html.close()
        if arguments['pdf']:
            pdffile = f'{output_path}{str(total)}-{(card_colour.upper().strip("#"))}.pdf'
            print_pdf(filename, pdffile)
        current_count = total
        total += 1
    if current_count == 1:
        print(f"{str(current_count)} card written")
    elif current_count > 1:
        print(f"{str(current_count)} cards written")


def convert_logo(logo):
    """When logo is chosen, will load, resize, then add the logo to the HTML"""
    file_name, file_ext = os.path.splitext(logo)
    basewidth = 40
    img = Image.open(logo)
    wpercent = (basewidth / float(img.size[0]))
    hsize = int((float(img.size[1]) * float(wpercent)))
    img = img.resize((basewidth, hsize), Image.Resampling.LANCZOS)
    resized_logo = f'{file_name}-resized{file_ext}'
    img.save(resized_logo)

    filetype = imghdr.what(resized_logo)
    file_handle = open(resized_logo, 'rb').read()
    b64_logo = base64.b64encode(file_handle).decode('utf-8')
    data_uri = f'data:image/{filetype};base64,{b64_logo}'

    return data_uri

def generate_numbers():
    """Generate random numbers for each column"""
    rand = Random()
    card_array = []
    for _ in range(5):
        b_nums = rand.sample(range(1, 16), 1)[0]
        while b_nums in card_array:
            b_nums = rand.sample(range(1, 16), 1)[0]
        card_array.append(b_nums)
        i_nums = rand.sample(range(16, 31), 1)[0]
        while i_nums in card_array:
            i_nums = rand.sample(range(16, 31), 1)[0]
        card_array.append(i_nums)
        n_nums = rand.sample(range(31, 46), 1)[0]
        while n_nums in card_array:
            n_nums = rand.sample(range(31, 46), 1)[0]
        card_array.append(n_nums)
        g_nums = rand.sample(range(46, 61), 1)[0]
        while g_nums in card_array:
            g_nums = rand.sample(range(46, 61), 1)[0]
        card_array.append(g_nums)
        o_nums = rand.sample(range(61, 76), 1)[0]
        while o_nums in card_array:
            o_nums = rand.sample(range(61, 76), 1)[0]
        card_array.append(o_nums)
    return card_array

def print_pdf(html_file, out_file):
    """Configure options for printing to PDF"""
    if sys.platform == 'linux':
        left = '0.25in'
        right = '0.25in'
    else:
        left = '0.1in'
        right = '0in'
    options = {
        'page-size': 'Letter',
        'page-width': '8.5in',
        'page-height': '11in',
        'orientation': 'Landscape',
        'margin-top': '0.5in',
        'margin-right': right,
        'margin-bottom': '0.25in',
        'margin-left': left,
        'quiet': ''
    }
    with open(html_file, "r") as html:
        html = html.read().replace(' - CLICK HERE TO CLEAR CARD', '')
        html = html.replace('<a href="https://github.com/digitalsleuth/bingo-card-generator" class="footer"></a>',
                            '<div align="center" style="font-family: Roboto Condensed">'
                            'https://github.com/digitalsleuth/bingo-card-generator</div>')
        html = html.replace('<select', '<!-- <select').replace('</select>', '</select> -->')

    html_back = f'{html_file}.html'
    with open(html_back, "w") as backup:
        backup.write(html)
    pdfkit.from_file(html_back, out_file, options=options)
    os.remove(html_back)

def grab_numbers(arguments):
    """Extracts the generated numbers from the HTML files to forward to other functions"""
    num = int(arguments['num'])
    if num == 1:
        amt = 'card'
    else:
        amt = 'cards'
    print(f"Extracting numbers from {num} {amt}")
    total = 1
    output_path = arguments['output']
    if arguments['everything'] and not arguments['base_colour']:
        basecolour = arguments['card_colour'].upper()
    else:
        basecolour = arguments['base_colour'].upper()
    pattern = '\$card\d = \[*;*'
    if '.html' not in basecolour:
        basecolour = basecolour + '.html'
    while total <= num:
        input_filename = f'{output_path}{os.sep}{str(total)}-{basecolour}'
        input_filename = input_filename.replace('#', '')
        input_file = open(input_filename, 'r')
        input_file = input_file.readlines()
        html_name, _ = os.path.splitext(basecolour)
        output_filename = f'{output_path}{os.sep}{str(total)}-{(html_name.upper())}.csv'
        output_file = open(output_filename, 'w+')
        full_sheet = []
        for line in input_file:
            match = re.match(pattern, line)
            if match:
                row_1, row_2, row_3, row_4, row_5, \
                row_6, row_7, row_8, row_9, row_10 = ([] for i in range(10))
                line = re.sub(pattern, '', line)
                line = line.replace("];\n", "").replace(',', '')
                line = line.split()
                for i in line:
                    full_sheet.append(i)

        indices = {
            1: [[0, 1, 2, 3, 4], [25, 26, 27, 28, 29], [50, 51, 52, 53, 54]],
            2: [[5, 6, 7, 8, 9], [30, 31, 32, 33, 34], [55, 56, 57, 58, 59]],
            3: [[10, 11, 12, 13, 14], [35, 36, 37, 38, 39], [60, 61, 62, 63, 64]],
            4: [[15, 16, 17, 18, 19], [40, 41, 42, 43, 44], [65, 66, 67, 68, 69]],
            5: [[20, 21, 22, 23, 24], [45, 46, 47, 48, 49], [70, 71, 72, 73, 74]],
            6: [[75, 76, 77, 78, 79], [100, 101, 102, 103, 104], [125, 126, 127, 128, 129]],
            7: [[80, 81, 82, 83, 84], [105, 106, 107, 108, 109], [130, 131, 132, 133, 134]],
            8: [[85, 86, 87, 88, 89], [110, 111, 112, 113, 114], [135, 136, 137, 138, 139]],
            9: [[90, 91, 92, 93, 94], [115, 116, 117, 118, 119], [140, 141, 142, 143, 144]],
            10: [[95, 96, 97, 98, 99], [120, 121, 122, 123, 124], [145, 146, 147, 148, 149]]
        }

        all_rows = [row_1, row_2, row_3, row_4, row_5, row_6, row_7, row_8, row_9, row_10]
        for idx in range(len(indices)):
            for i in indices.get((idx + 1)):
                (all_rows[idx]).extend(map(full_sheet.__getitem__, i))
            all_rows[idx].insert(5, " ")
            all_rows[idx].insert(11, " ")
        all_rows[2][2] = all_rows[2][8] = all_rows[2][14] = "*"
        all_rows[7][2] = all_rows[7][8] = all_rows[7][14] = "*"
        all_rows.insert(5, [", "])
        for row in all_rows:
            row = ','.join(map(str, row)) + '\n'
            output_file.write(row)
        output_file.close()
        total += 1
    print("Extraction Complete")

def generate_excel(number_of_csvs, base_filename, excel_name, source_path):
    """Takes the bingo numbers and card numbers and writes the data to an Excel file"""
    header = [' ', 'B', 'I', 'N', 'G', 'O',
              ' ', 'B', 'I', 'N', 'G', 'O',
              ' ', 'B', 'I', 'N', 'G', 'O']
    excel_name = f'{source_path}{os.sep}{(excel_name.upper())}'
    call_sheet_header = ['B', 'I', 'N', 'G', 'O']
    call_sheet = NamedStyle(name="call_sheet")
    call_sheet.alignment.horizontal = 'center'
    call_sheet.alignment.vertical = 'center'
    bingo_header = NamedStyle(name="bingo_header")
    bingo_header.font = Font(bold=True, name='Arial', size='15')
    bingo_header.alignment.horizontal = 'center'
    bingo_header.alignment.vertical = 'center'
    called_number = PatternFill(bgColor="FFC000")
    free_space = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    borders = PatternFill(start_color='B2B2B2', end_color='B2B2B2', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')
    writer = Workbook(excel_name)
    call_worksheet = writer.create_sheet('CALL')
    call_worksheet.append(call_sheet_header)
    base_name, _ = os.path.splitext(base_filename)
    for csvnum in range(1, number_of_csvs + 1):
        csvfile = (f'{source_path}{os.sep}{str(csvnum)}-{(base_name.upper())}.csv')
        worksheet = writer.create_sheet(str(csvnum))
        readcsv = open(csvfile, 'r', newline='', encoding='utf-8')
        reader = csv.reader(readcsv)
        for row in reader:
            for item in row:
                try:
                    row[row.index(item)] = int(row[row.index(item)])
                except ValueError:
                    pass
            worksheet.append(row)
        readcsv.close()
        os.remove(csvfile)
    writer.save(excel_name)
    wb = load_workbook(excel_name)
    wb.add_named_style(bingo_header)
    wb.add_named_style(call_sheet)
    ws_call = wb['CALL']
    for row in range(1, 17):
        ws_call.row_dimensions[row].height = 20
    call_font = Font(bold=True, name='Arial', size='20')
    col_b = ws_call.column_dimensions['A']
    col_i = ws_call.column_dimensions['B']
    col_n = ws_call.column_dimensions['C']
    col_g = ws_call.column_dimensions['D']
    col_o = ws_call.column_dimensions['E']
    for col_letter in (col_b, col_i, col_n, col_g, col_o):
        col_letter.font = call_font
        col_letter.alignment = alignment
    for _, cell in enumerate(ws_call["A1":"E1"]):
        for _, cell_obj in enumerate(cell):
            cell_obj.style = call_sheet
            cell_obj.font = Font(bold=True, name='Arial', size='20', color='FF0000')
    for sheet in range(1, number_of_csvs + 1):
        ws = wb[str(sheet)]
        ws.conditional_formatting.add('B1:B14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(B1,CALL!$A$2:$A$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('C1:C14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(C1,CALL!$B$2:$B$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('D1:D14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(D1,CALL!$C$2:$C$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('E1:E14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(E1,CALL!$D$2:$D$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('F1:F14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(F1,CALL!$E$2:$E$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('H1:H14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(H1,CALL!$A$2:$A$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('I1:I14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(I1,CALL!$B$2:$B$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('J1:J14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(J1,CALL!$C$2:$C$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('K1:K14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(K1,CALL!$D$2:$D$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('L1:L14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(L1,CALL!$E$2:$E$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('N1:N14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(N1,CALL!$A$2:$A$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('O1:O14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(O1,CALL!$B$2:$B$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('P1:P14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(P1,CALL!$C$2:$C$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('Q1:Q14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(Q1,CALL!$D$2:$D$16,1,FALSE)))'], fill=called_number))
        ws.conditional_formatting.add('R1:R14', FormulaRule(formula=['NOT(ISNA(VLOOKUP(R1,CALL!$E$2:$E$16,1,FALSE)))'], fill=called_number))
        for row in range(1, 18):
            ws.row_dimensions[row].height = 20
        column = 1
        while column <= 19:
            col = get_column_letter(column)
            ws.column_dimensions[col].width = 5
            column += 1
        ws.insert_cols(1)
        ws.insert_rows(0)
        ws.insert_rows(1)
        ws.insert_rows(8)
        ws['D5'].fill = ws['J5'].fill = ws['P5'].fill = free_space
        ws['D12'].fill = ws['J12'].fill = ws['P12'].fill = free_space
        ws['D5'].alignment = ws['J5'].alignment = ws['P5'].alignment = alignment
        ws['D12'].alignment = ws['J12'].alignment = ws['P12'].alignment = alignment
        for col, val in enumerate(header, start=1):
            ws.cell(row=2, column=col).value = val
            ws.cell(row=9, column=col).value = val
        for cell in ws["1:1"]:
            cell.fill = borders
        for cell in ws["2:2"]:
            cell.style = bingo_header
        for cell in ws["8:8"]:
            cell.fill = borders
        for cell in ws["9:9"]:
            cell.style = bingo_header
        for cell in ws["15:15"]:
            cell.fill = borders
        for _, cell in enumerate(ws["A1":"A15"]):
            for _, cell_obj in enumerate(cell):
                cell_obj.fill = borders
        for _, cell in enumerate(ws["G1":"G15"]):
            for _, cell_obj in enumerate(cell):
                cell_obj.fill = borders
        for _, cell in enumerate(ws["M1":"M15"]):
            for _, cell_obj in enumerate(cell):
                cell_obj.fill = borders
        for _, cell in enumerate(ws["S1":"S15"]):
            for _, cell_obj in enumerate(cell):
                cell_obj.fill = borders
        for row in ws.rows:
            for cell in row:
                cell.alignment = alignment
    wb.save(excel_name)

def main():
    """Parse arguments for PDF, card and dauber colour, and dauber shape"""
    arg_parse = argparse.ArgumentParser(
        description='Bingo Card Generator v' + str(__version__),
        epilog=f"If you'd like to see a few other color options, you can visit:\n{__colour_groups__}",
        formatter_class=argparse.RawTextHelpFormatter)
    arg_parse.add_argument('-v', '--version',
                           action='version',
                           version='%(prog)s ' + str(__version__))
    group = arg_parse.add_argument_group(title='positional arguments',
                                         description='''NUM_OF_CARDS

Customization options:

-e, --everything              Creates HTML files, PDF's and spreadsheet - use -c, -d, and -s for desired customization
                              otherwise your cards will be set to default values.
-p, --pdf                     Convert generated HTML files to PDFs
-o, --output <output_dir>     Choose output directory
-c, --card-colour <colour>    Colour for the card - default is BLUE
-d, --dauber-colour <colour>  Colour for the dauber - default is RED
-s, --dauber-shape <shape>    Shape for the dauber - default is CIRCLE
                              Options are: checkmark, circle, clover, heart, logo, maple-leaf, moon, square, star, unicorn, x-mark
-l, --logo <image_file>       If logo is chosen, this option points to a JPG or PNG to use as a dauber
-a, --allow-select            If chosen, provides a dropdown box on the HTML card for the player to change their dauber at will
-t, --title                   Gives you the option to add a title to the card, eg "Bonanza Bingo!"
-z, --easy                    Enables "Easy Mode" - meaning once a number is clicked on one card, it is selected on all cards
-x, --excel <path>            When not choosing "everything", generates an Excel document for the call numbers - requires -b
-b, --base-colour <colour>    Identifies the base colour of the HTML files from which to generate the Excel Spreadsheet
''')
    group.add_argument('-p', '--pdf', action='store_true', help=argparse.SUPPRESS)
    group.add_argument('-o', '--output', metavar='', help=argparse.SUPPRESS, required=True)
    group.add_argument('num', metavar='', help=argparse.SUPPRESS, type=int, nargs=1)
    group.add_argument('-c', '--card-colour', help=argparse.SUPPRESS, default='blue')
    group.add_argument('-d', '--dauber-colour', help=argparse.SUPPRESS, default='red')
    group.add_argument('-s', '--dauber-shape', help=argparse.SUPPRESS,
                       choices=['square',
                                'circle',
                                'maple-leaf',
                                'heart',
                                'star',
                                'moon',
                                'unicorn',
                                'clover',
                                'logo',
                                'checkmark',
                                'x-mark'], default='circle')
    group.add_argument('-l', '--logo', help=argparse.SUPPRESS)
    group.add_argument('-a', '--allow-select', help=argparse.SUPPRESS, action='store_true')
    group.add_argument('-t', '--title', help=argparse.SUPPRESS, default='')
    group.add_argument('-b', '--base-colour', help=argparse.SUPPRESS)
    group.add_argument('-x', '--excel', help=argparse.SUPPRESS)
    group.add_argument('-e', '--everything', help=argparse.SUPPRESS, action='store_true')
    group.add_argument('-z', '--easy', help=argparse.SUPPRESS, action='store_true')

    if len(sys.argv[1:]) == 0:
        arg_parse.print_help()
        arg_parse.exit()

    args = arg_parse.parse_args()
    all_args = vars(args)
    all_args['num'] = all_args['num'][0]
    if all_args['excel'] and all_args['base_colour']:
        grab_numbers(all_args)
        generate_excel(all_args['num'], all_args['base_colour'],
                       all_args['excel'], all_args['output'])
    elif all_args['excel'] and not all_args['base_colour']:
        print("The Excel option requires the -b, --base-colour value as well")
        raise SystemExit(0)
    elif all_args['everything'] and all_args['excel']:
        all_args['pdf'] = True
        create_card(all_args)
        grab_numbers(all_args)
        generate_excel(all_args['num'], all_args['card_colour'],
                       all_args['excel'], all_args['output'])
    elif all_args['everything'] and not all_args['excel']:
        all_args['pdf'] = True
        create_card(all_args)
        grab_numbers(all_args)
        generate_excel(all_args['num'], all_args['card_colour'],
                       str(f'{all_args["card_colour"]}-cards.xlsx'), all_args['output'])
    else:
        create_card(all_args)


if __name__ == '__main__':
    main()
