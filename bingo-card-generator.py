#!/usr/bin/env python3
'''
Tool to generate a clickable bingo card for virtual bingo events and to
generate PDF's for printing these cards.

Requires pdfkit and wkhtmltopdf for OS it's being run on
(Windows requires an exe, linux install from pkg manager)

This script currently will only generate 6 cards on one sheet (2 rows of 3 cards).
Hex codes can be used in place of words for colours,
but must be wrapped in quotes on the command line

While not normally necessary for CSS, the float values are required for
the proper printing of the PDF's with wkhtmltopdf.
If these are removed, you can expect the PDF's to be off-center or misaligned.

CSS Maple Leaf author Andre Lopes - https://codepen.io/alldrops/pen/jAzZmw
'''

import argparse
from random import Random
import pdfkit
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle, PatternFill

__author__ = 'Corey Forman'
__date__ = '24 July 2021'
__version__ = '1.3.1'
__description__ = 'Interactive Bingo Card Generator'


def createCard(arguments):
    """Creates the HTML version of the card"""
    card_colour = arguments['card_colour']
    dauber_colour = arguments['dauber_colour']
    dauber_shape = arguments['dauber_shape']
    total = 1
    head1 = '''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
'''

    head2 = '''<style>
.clear { width: 100%; clear: both; }
.card { margin-top: 25px; float: left; width: 400px; height: auto; background-color: ''' + card_colour + '''; border-radius: 5px; padding: 10px; }
.clear-card { position: absolute; }
.card-number { position: relative; justify-items: center; text-align: center; font-family: "Roboto Condensed", sans-serif; font-weight: bold; color: ''' + card_colour + '''; }
.grid-container { float: center; display: grid; grid-template-columns: auto auto auto; grid-gap: 5px; grid-template-rows: auto; grid-row-gap: 0px; justify-items: center; align-items: center; }
.grid-child { float: left; justify-items: center; align-items: center; padding: 5px; }
.center { display: flex; justify-content: center; position: absolute;  }
.headers > div { float: left; width: 80px; text-align: center; }
.headers > div span { font-size: 30px; color: #fff; font-weight: bold; font-family: Arial, Helvetica, sans-serif;}
.column { float: left; width: 80px; text-align: center; }
.number { padding: 20px 0; border: 2px solid ''' + card_colour + '''; background-color: #fff; font-family: "Roboto Condensed", sans-serif; font-weight: normal; height: 16px; }
.number span { color: #000; font-size: 20px; }
.number span:hover { text-shadow: 0 0 5px rgba(0,0,0,0.5); }
.circle-dauber { width: 30px; height: 30px; background: ''' + dauber_colour + '''; border-radius: 50%; position: relative; top: 10%; left: 30%; padding: 0px; box-shadow: inset 0.1em 0.1em 0.1em 0 rgba(255,255,255,0.5), inset -0.1em -0.1em 0.1em 0 rgba(0,0,0,0.5); -ms-transform: translateY(-50%); transform: translateY(-30%); }
.square-dauber { width: 30px; height: 30px; background: ''' + dauber_colour + '''; border-radius: 10%; position: relative; top: 36%; left: 30%; padding: 0px; box-shadow: inset 0.1em 0.1em 0.1em 0 rgba(255,255,255,0.5), inset -0.1em -0.1em 0.1em 0 rgba(0,0,0,0.5); -ms-transform: translateY(-50%); transform: translateY(-40%); }
.maple-dauber { display: flex; align-items: center; justify-content: center; position: relative; margin-left: 0px; margin-right: 0px; margin-top: 11px; margin-bottom: 12px; }
.maple-dauber:after { position: absolute; content: ""; width: 40px; height: 40px; background: #F00; -ms-transform: translateY(-20%); transform: translateY(-5%);
    clip-path: polygon(47% 100%, 48% 70%, 25% 73%, 28% 65%, 7% 47%, 11% 44%, 8%     30%, 20% 32%, 23% 27%, 35% 40%, 32% 13%, 39% 16%, 50% 0, 61% 16%, 68% 13%,     65% 40%, 77% 27%, 80% 32%, 92% 30%, 89% 44%, 93% 47%, 72% 65%, 75% 73%, 52% 70%, 53% 100%); }
</style>
</head>
<body>
'''

    body = '''
<div class="grid-container 1">

<div class="grid-child 1">
<div class="clear"></div>
<div class="card 1">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card1-c1"></div>
    <div class="number col-6" id="card1-c6"></div>
    <div class="number col-11" id="card1-c11"></div>
    <div class="number col-16" id="card1-c16"></div>
    <div class="number col-21" id="card1-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card1-c2"></div>
    <div class="number col-7" id="card1-c7"></div>
    <div class="number col-12" id="card1-c12"></div>
    <div class="number col-17" id="card1-c17"></div>
    <div class="number col-22" id="card1-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card1-c3"></div>
    <div class="number col-8" id="card1-c8"></div>
    <div class="number col-13" id="card1-c13"></div>
    <div class="number col-18" id="card1-c18"></div>
    <div class="number col-23" id="card1-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card1-c4"></div>
    <div class="number col-9" id="card1-c9"></div>
    <div class="number col-14" id="card1-c14"></div>
    <div class="number col-19" id="card1-c19"></div>
    <div class="number col-24" id="card1-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card1-c5"></div>
    <div class="number col-10" id="card1-c10"></div>
    <div class="number col-15" id="card1-c15"></div>
    <div class="number col-20" id="card1-c20"></div>
    <div class="number col-25" id="card1-c25"></div>
  </div>
</div>
</div>

<div class="grid-child 2">
<div class="clear"></div>
<div class="card 2">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card2-c1"></div>
    <div class="number col-6" id="card2-c6"></div>
    <div class="number col-11" id="card2-c11"></div>
    <div class="number col-16" id="card2-c16"></div>
    <div class="number col-21" id="card2-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card2-c2"></div>
    <div class="number col-7" id="card2-c7"></div>
    <div class="number col-12" id="card2-c12"></div>
    <div class="number col-17" id="card2-c17"></div>
    <div class="number col-22" id="card2-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card2-c3"></div>
    <div class="number col-8" id="card2-c8"></div>
    <div class="number col-13" id="card2-c13"></div>
    <div class="number col-18" id="card2-c18"></div>
    <div class="number col-23" id="card2-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card2-c4"></div>
    <div class="number col-9" id="card2-c9"></div>
    <div class="number col-14" id="card2-c14"></div>
    <div class="number col-19" id="card2-c19"></div>
    <div class="number col-24" id="card2-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card2-c5"></div>
    <div class="number col-10" id="card2-c10"></div>
    <div class="number col-15" id="card2-c15"></div>
    <div class="number col-20" id="card2-c20"></div>
    <div class="number col-25" id="card2-c25"></div>
  </div>
</div>
</div>

<div class="grid-child 3">
<div class="clear"></div>
<div class="card 3">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card3-c1"></div>
    <div class="number col-6" id="card3-c6"></div>
    <div class="number col-11" id="card3-c11"></div>
    <div class="number col-16" id="card3-c16"></div>
    <div class="number col-21" id="card3-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card3-c2"></div>
    <div class="number col-7" id="card3-c7"></div>
    <div class="number col-12" id="card3-c12"></div>
    <div class="number col-17" id="card3-c17"></div>
    <div class="number col-22" id="card3-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card3-c3"></div>
    <div class="number col-8" id="card3-c8"></div>
    <div class="number col-13" id="card3-c13"></div>
    <div class="number col-18" id="card3-c18"></div>
    <div class="number col-23" id="card3-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card3-c4"></div>
    <div class="number col-9" id="card3-c9"></div>
    <div class="number col-14" id="card3-c14"></div>
    <div class="number col-19" id="card3-c19"></div>
    <div class="number col-24" id="card3-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card3-c5"></div>
    <div class="number col-10" id="card3-c10"></div>
    <div class="number col-15" id="card3-c15"></div>
    <div class="number col-20" id="card3-c20"></div>
    <div class="number col-25" id="card3-c25"></div>
  </div>
</div>
</div>

</div>

<div class="grid-container 2">

<div class="grid-child 4">
<div class="clear"></div>
<div class="card 4">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card4-c1"></div>
    <div class="number col-6" id="card4-c6"></div>
    <div class="number col-11" id="card4-c11"></div>
    <div class="number col-16" id="card4-c16"></div>
    <div class="number col-21" id="card4-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card4-c2"></div>
    <div class="number col-7" id="card4-c7"></div>
    <div class="number col-12" id="card4-c12"></div>
    <div class="number col-17" id="card4-c17"></div>
    <div class="number col-22" id="card4-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card4-c3"></div>
    <div class="number col-8" id="card4-c8"></div>
    <div class="number col-13" id="card4-c13"></div>
    <div class="number col-18" id="card4-c18"></div>
    <div class="number col-23" id="card4-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card4-c4"></div>
    <div class="number col-9" id="card4-c9"></div>
    <div class="number col-14" id="card4-c14"></div>
    <div class="number col-19" id="card4-c19"></div>
    <div class="number col-24" id="card4-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card4-c5"></div>
    <div class="number col-10" id="card4-c10"></div>
    <div class="number col-15" id="card4-c15"></div>
    <div class="number col-20" id="card4-c20"></div>
    <div class="number col-25" id="card4-c25"></div>
  </div>
</div>
</div>

<div class="grid-child 5">
<div class="clear"></div>
<div class="card 5">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card5-c1"></div>
    <div class="number col-6" id="card5-c6"></div>
    <div class="number col-11" id="card5-c11"></div>
    <div class="number col-16" id="card5-c16"></div>
    <div class="number col-21" id="card5-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card5-c2"></div>
    <div class="number col-7" id="card5-c7"></div>
    <div class="number col-12" id="card5-c12"></div>
    <div class="number col-17" id="card5-c17"></div>
    <div class="number col-22" id="card5-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card5-c3"></div>
    <div class="number col-8" id="card5-c8"></div>
    <div class="number col-13" id="card5-c13"></div>
    <div class="number col-18" id="card5-c18"></div>
    <div class="number col-23" id="card5-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card5-c4"></div>
    <div class="number col-9" id="card5-c9"></div>
    <div class="number col-14" id="card5-c14"></div>
    <div class="number col-19" id="card5-c19"></div>
    <div class="number col-24" id="card5-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card5-c5"></div>
    <div class="number col-10" id="card5-c10"></div>
    <div class="number col-15" id="card5-c15"></div>
    <div class="number col-20" id="card5-c20"></div>
    <div class="number col-25" id="card5-c25"></div>
  </div>
</div>
</div>

<div class="grid-child 6">
<div class="clear"></div>
<div class="card 6">
  <div class="headers">
    <div><span>B</span></div>
    <div><span>I</span></div>
    <div><span>N</span></div>
    <div><span>G</span></div>
    <div><span>O</span></div>
  </div>
  <div class="column 1">
    <div class="number col-1" id="card6-c1"></div>
    <div class="number col-6" id="card6-c6"></div>
    <div class="number col-11" id="card6-c11"></div>
    <div class="number col-16" id="card6-c16"></div>
    <div class="number col-21" id="card6-c21"></div>
  </div>
  <div class="column 2">
    <div class="number col-2" id="card6-c2"></div>
    <div class="number col-7" id="card6-c7"></div>
    <div class="number col-12" id="card6-c12"></div>
    <div class="number col-17" id="card6-c17"></div>
    <div class="number col-22" id="card6-c22"></div>
  </div>
  <div class="column 3">
    <div class="number col-3" id="card6-c3"></div>
    <div class="number col-8" id="card6-c8"></div>
    <div class="number col-13" id="card6-c13"></div>
    <div class="number col-18" id="card6-c18"></div>
    <div class="number col-23" id="card6-c23"></div>
  </div>
  <div class="column 4">
    <div class="number col-4" id="card6-c4"></div>
    <div class="number col-9" id="card6-c9"></div>
    <div class="number col-14" id="card6-c14"></div>
    <div class="number col-19" id="card6-c19"></div>
    <div class="number col-24" id="card6-c24"></div>
  </div>
  <div class="column 5">
    <div class="number col-5" id="card6-c5"></div>
    <div class="number col-10" id="card6-c10"></div>
    <div class="number col-15" id="card6-c15"></div>
    <div class="number col-20" id="card6-c20"></div>
    <div class="number col-25" id="card6-c25"></div>
  </div>
</div>
</div>

</div>

<script src='https://cdnjs.cloudflare.com/ajax/libs/jquery/3.6.0/jquery.min.js'></script>
<script id='rendered-js' >
'''

    js1 = '''
$i = [$card1, $card2, $card3, $card4, $card5, $card6];

$(document).ready(function() {
  for ($card = 1; $card <= 6; $card++) {
    for ($x = 0; $x <= 24; $x++) {
    $("<span>" + $i[($card - 1)][$x] + "</span>").appendTo('#card' + $card + '-c' + ($x + 1));
      }
  }
'''
    js2 = '''
  $('.number').click(function() {
'''
    js3 = '''
  $('#clear-card').click(function() {
    location.reload();
    });
  });
});
</script>
</body>
</html>
'''

    while total <= int(arguments['num'][0]):
        filename = str(total) + '-' + card_colour.strip('#') + '.html'
        html = open(filename, 'w')
        html.write(head1)
        html.write("<title>CARD " + str(total) + "</title>")
        html.write(head2)
        count = 1
        card_clear = '<div class="card-number" id="clear-card">CARD ' + str(total) + ' - CLICK HERE TO CLEAR CARD</div>'
        free_space = '$(\".col-13\").html(\'<span style=\"color:' + card_colour + '; font-weight:bold\">FREE</span>\');'
        dauber_script = '$(this).html(\'<div class=\"' + dauber_shape + '-dauber\"></div>\');'
        html.write(card_clear)
        html.write(body)
        while count <= 6:
            nums = str(genNums())
            html.write('$card' + str(count) + ' = ' + nums + ';\n')
            count += 1
        html.write(js1)
        html.write(free_space)
        html.write(js2)
        html.write(dauber_script)
        html.write(js3)
        html.close()
        if arguments['pdf']:
            pdffile = str(total) + '-' + card_colour.strip('#') + '.pdf'
            pdfPrint(filename, pdffile)
        current_count = total
        total += 1
    print("{} cards written.".format(str(current_count)))


def genNums():
    """Generate random numbers for each column"""
    rand = Random()
    card_array = []
    for _ in range(5):
        b = rand.sample(range(1, 16), 1)[0]
        while b in card_array:
            b = rand.sample(range(1, 16), 1)[0]
        card_array.append(b)
        i = rand.sample(range(16, 31), 1)[0]
        while i in card_array:
            i = rand.sample(range(16, 31), 1)[0]
        card_array.append(i)
        n = rand.sample(range(31, 46), 1)[0]
        while n in card_array:
            n = rand.sample(range(31, 46), 1)[0]
        card_array.append(n)
        g = rand.sample(range(46, 61), 1)[0]
        while g in card_array:
            g = rand.sample(range(46, 61), 1)[0]
        card_array.append(g)
        o = rand.sample(range(61, 76), 1)[0]
        while o in card_array:
            o = rand.sample(range(61, 76), 1)[0]
        card_array.append(o)
    return card_array


def pdfPrint(html_file, out_file):
    """Configure options for printing to PDF"""
    options = {
        'page-size': 'Letter',
        'page-width': '8.5in',
        'page-height': '11in',
        'orientation': 'Landscape',
        'margin-top': '0.5in',
        'margin-right': '0in',
        'margin-bottom': '0.5in',
        'margin-left': '0in',
        'quiet': ''
    }
    pdfkit.from_file(html_file, out_file, options=options)

def grabNumbers(arguments):
    num = int(arguments['num'][0])
    total = 1
    basename = arguments['basename']
    pattern = '\$card\d = \[*;*'
    if '.html' not in basename:
        basename = basename + '.html'
    while total <= num:
        input_filename = str(total) + "-" + basename
        input_file = open(input_filename, 'r')
        input_file = input_file.readlines()
        output_filename = str(total) + "-" + basename.strip('.html') + '.csv'
        output_file = open(output_filename, 'w+')
        full_sheet = []
        for line in input_file:
            match = re.match(pattern, line)
            if match:
                row_1, row_2, row_3, row_4, row_5, \
                row_6, row_7, row_8, row_9, row_10 = ([] for i in range(10))
                line = re.sub(pattern, '', line)
                line = line.replace("];\n", "").replace(',','')
                line = line.split()
                for i in line:
                    full_sheet.append(i)

        indices = {
            1: [[0, 1, 2, 3, 4], [25, 26, 27, 28, 29], [50, 51, 52, 53, 54]],
            2: [[5, 6, 7, 8, 9], [30, 31, 32, 33, 34], [55, 56, 57, 58, 59]],
            3: [[10, 11, 12, 13, 14], [35, 36, 37, 38, 39], [60, 61, 62, 63, 64]],
            4: [[15, 16, 17, 18, 19], [40, 41, 42, 43, 44], [65, 66, 67, 68, 69]],
            5: [[20, 21, 22, 23, 24], [45, 46, 47, 48, 49], [70, 71, 72, 73, 74]],
            6: [[75, 76, 77, 78, 79], [100, 101, 102, 103, 104], [125, 126, 127, 128, 129]],
            7: [[80, 81, 82, 83, 84], [105, 106, 107, 108, 109], [130, 131, 132, 133, 134]],
            8: [[85, 86, 87, 88, 89], [110, 111, 112, 113, 114], [135, 136, 137, 138, 139]],
            9: [[90, 91, 92, 93, 94], [115, 116, 117, 118, 119], [140, 141, 142, 143, 144]],
            10: [[95, 96, 97, 98, 99], [120, 121, 122, 123, 124], [145, 146, 147, 148, 149]]
        }

        all_rows = [row_1, row_2, row_3, row_4, row_5, row_6, row_7, row_8, row_9, row_10]
        for idx in range(len(indices)):
            for i in indices.get((idx + 1)):
                (all_rows[idx]).extend(map(full_sheet.__getitem__, i))
            all_rows[idx].insert(5, " ")
            all_rows[idx].insert(11, " ")
        all_rows[2][2] = all_rows[2][8] = all_rows[2][14] = all_rows[7][2] = all_rows[7][8] = all_rows[7][14] = "*"
        all_rows.insert(5, [", "])
        for row in all_rows:
            row = ','.join(map(str, row)) + '\n'
            output_file.write(row)
        output_file.close()
        total += 1
    print("Complete")

def writeToExcel(number_of_csvs, base_filename, excel_name):
    header = [' ', 'B', 'I', 'N', 'G', 'O', ' ', 'B', 'I', 'N', 'G', 'O', ' ', 'B', 'I', 'N', 'G', 'O']
    bingo_header = NamedStyle(name="bingo_header")
    bingo_header.font = Font(bold=True, name='Arial', size='15')
    bingo_header.alignment.horizontal = 'center'
    bingo_header.fill.start_color="FFFFFF"
    bingo_header.fill.end_color="FFFFFF"
    bingo_header.fill.fill_type="solid"
    free_space = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    border = NamedStyle(name="border")
    border.fill.start_color="B2B2B2"
    border.fill.end_color="B2B2B2"
    border.fill.fill_type="solid"
    border_columns = PatternFill(start_color="B2B2B2", end_color="B2B2B2", fill_type="solid")
    writer = pd.ExcelWriter(excel_name)
    for csvnum in range(1, number_of_csvs + 1):
        csvfile = str(csvnum) + '-' + base_filename.strip('.html') + '.csv'
        df = pd.read_csv(csvfile, sep=',', header=None)
        df.to_excel(writer, sheet_name=str(csvnum), index=False, columns=None, startrow=1, startcol=1)
        workbook = writer.book
        worksheet = writer.sheets[str(csvnum)]
        format = workbook.add_format({"font_name": "Arial", "font_size": "13", "align": "center", "valign": "vcenter"})
        format.set_align('center')
        worksheet.set_column('A:S', 8, format)
        worksheet.set_default_row(30)
    writer.save()
    wb = load_workbook(excel_name)
    wb.add_named_style(bingo_header)
    wb.add_named_style(border)
    for sheet in range(1, number_of_csvs + 1):
        ws = wb[str(sheet)]
        ws.delete_rows(2)
        ws.insert_rows(1)
        ws.insert_rows(8)
        ws['D5'].fill = ws['J5'].fill = ws['P5'].fill = ws['D12'].fill = ws['J12'].fill = ws['P12'].fill = free_space
        for col, val in enumerate(header, start=1):
            ws.cell(row=2, column=col).value = val
            ws.cell(row=9, column=col).value = val
        for cell in ws["1:1"]:
            cell.style = 'border'
        for cell in ws["2:2"]:
            cell.style = 'bingo_header'
        for cell in ws["8:8"]:
            cell.style = 'border'
        for cell in ws["9:9"]:
            cell.style = 'bingo_header'
        for cell in ws["15:15"]:
            cell.style = 'border'
        for i, cell in enumerate(ws["A1":"A15"]):
            for n, cellObj in enumerate(cell):
                cellObj.style = border
        for i, cell in enumerate(ws["G1":"G15"]):
            for n, cellObj in enumerate(cell):
                cellObj.style = border
        for i, cell in enumerate(ws["M1":"M15"]):
            for n, cellObj in enumerate(cell):
                cellObj.style = border
        for i, cell in enumerate(ws["S1":"S15"]):
            for n, cellObj in enumerate(cell):
                cellObj.style = border
    wb.save(excel_name)

def main():
    """Parse arguments for PDF, card and dauber colour, and dauber shape"""
    arg_parse = argparse.ArgumentParser(description='Interactive Bingo Card Generator and PDF converter v' + str(__version__), formatter_class=argparse.RawTextHelpFormatter)
    arg_parse.add_argument('num', metavar='<# of cards>', help='Number of cards to generate', type=int, nargs=1)
    arg_parse.add_argument('-p', '--pdf', action='store_true', help='Convert the generated HTML file to a PDF')
    arg_parse.add_argument('-c', '--card-colour', help='Colour for the card', default='#1644b9')
    arg_parse.add_argument('-d', '--dauber-colour', help='Colour for the dauber', default='red')
    arg_parse.add_argument('-s', '--dauber-shape', help='Shape of the dauber [square|circle|maple]', default='circle')
    arg_parse.add_argument('-b', '--basename', help='Basename for the file (ie: 1-blue.html = blue.html, - is implied)')
    arg_parse.add_argument('-x', '--excel', metavar='filename', help='Output all CSVs to a single Excel Workbook (Macro Enabled)')

    args = arg_parse.parse_args()
    all_args = vars(args)

    if all_args['excel']:
        grabNumbers(all_args)
        writeToExcel(all_args['num'][0], all_args['basename'], all_args['excel'])
    else:
        createCard(all_args)

if __name__ == '__main__':
    main()
